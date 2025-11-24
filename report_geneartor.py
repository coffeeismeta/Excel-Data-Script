import openpyxl
from openpyxl.styles import Font

#1. Load existing data

wb = openpyxl.load_workbook("data.xlsx")
sheet = wb.active
amounts = []

#Loop through each row (skip header)

#loop thorugh each row (skip header)

for row in range(2, sheet.max_row + 1):
    value = sheet.cell(row=row, column=2).value
    if isinstance(value, (int, float)):
        amounts.append(value)

# 2. Summary Calculations ##

total = sum(amounts)
average = total/ len(amounts)
maximum = max(amounts)
minimum = min(amounts)

# 3. Create report workbook

report_wb = openpyxl.Workbook()
report_sheet = report_wb.active
report_sheet.title = "Summary Report"

# 4. Write Summary to Report

report_sheet["A1"] = "Metric"
report_sheet["B1"] = "Value"

# Bold Header
header_font = Font(bold=True)
report_sheet["A1"].font = header_font
report_sheet["B1"].font = header_font

data = [
    ("Total", total),
    ("Average", average),
    ("Maximum", maximum),
]

for i, (label, value) in enumerate(data, start=2):
    report_sheet[f"A{i}"] = label
    report_sheet[f"B{i}"] = value

# Auto-adjust column width

for col in ["A", "B"]:
    report_sheet.column_dimensions[col].width = 15


# 5. Save the Report

report_wb.save("report.xlsx")
print("Report Generated: report.xlsx")
